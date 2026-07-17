"""Tests de la exportacion a Excel.

Fijan las definiciones de negocio de las columnas consolidadas ("Total WIP",
"Total completos") y la estructura de encabezados de dos niveles.
"""

from __future__ import annotations

import io
from datetime import date

from openpyxl import load_workbook

from rbom_api.domain.modelo import ArbolPT, DemandaPT, NodoComponente, PasoRuta
from rbom_api.services.export_service import (
    _total_completos,
    _total_wip,
    exportar_arboles,
)


CORTE, DOBLEZ, EMBARQUES, ALM_WIP = 3, 4, 13, 16


def _paso(idProceso: int, proceso: str, wip: float, **kw) -> PasoRuta:
    return PasoRuta(
        orden=kw.get("orden", 1),
        idProceso=idProceso,
        proceso=proceso,
        es_virtual=kw.get("es_virtual", False),
        wip_en_paso=wip,
        disponibles=kw.get("disponibles", 0.0),
        recibidas=kw.get("recibidas", 0.0),
        liberadas=kw.get("liberadas", 0.0),
        req_paso=kw.get("req_paso", 0.0),
    )


def _comp(clave: str, ruta: list[PasoRuta], **kw) -> NodoComponente:
    return NodoComponente(
        idComp=kw.get("idComp", 1),
        clave=clave,
        descripcion=kw.get("descripcion", "desc"),
        nivel=kw.get("nivel", 2),
        tipo_material=kw.get("tipo_material", 3),
        cantidad_ensamble_total=kw.get("cant", 1.0),
        req_bruto=kw.get("req_bruto", 100.0),
        wip_total=kw.get("wip_total", 0.0),
        req_neto=kw.get("req_neto", 0.0),
        ruta=ruta,
    )


def _arbol(componentes: list[NodoComponente], pt: str = "PT-1") -> ArbolPT:
    return ArbolPT(
        pt=DemandaPT(
            idMaterial=1, PT=pt, Descripcion="Producto",
            Cliente="Cliente X", Ciudad="Ciudad Y",
            PiezasPend=100, FechaPromMin=date(2026, 5, 1),
            FechaPromMax=date(2026, 6, 1), PiezasPastDue=10,
        ),
        componentes=componentes,
    )


# ---- Definiciones de negocio ------------------------------------------------

def test_total_completos_es_el_wip_del_ultimo_paso_de_la_ruta():
    """Completas = las que esperan en el ultimo paso (ya pasaron todo lo previo).

    `wip_en_paso[X]` = piezas esperando ENTRAR a X (trampa #10 del contrato).
    Las 10 de Corte y las 3 de Doblez todavia no completan su ruta; las 7 del
    buffer si.
    """
    comp = _comp("C1", [
        _paso(CORTE, "Corte", 10),
        _paso(DOBLEZ, "Doblez", 3),
        _paso(ALM_WIP, "Almacen WIP", 7, es_virtual=True),
    ])
    assert _total_completos(comp) == 7


def test_total_completos_no_es_el_wip_del_ultimo_proceso_real():
    """Regresion del caso canonico 91711066-RA (verificado contra BD real).

    Tomar el WIP del ultimo proceso de fabricacion invierte la realidad:
      - 90358715-RA: 4 pzs esperando Doblez -> 0 completas (les falta Doblez),
        pero el ultimo proceso real diria 4.
      - 91711040-RA: 9 pzs en Almacen WIP  -> 9 completas,
        pero el ultimo proceso real (Doblez=0) diria 0.
    """
    c1 = _comp("90358715-RA", [
        _paso(CORTE, "Corte", 0),
        _paso(DOBLEZ, "Doblez", 4),
        _paso(ALM_WIP, "Almacen WIP", 0, es_virtual=True),
    ])
    c2 = _comp("91711040-RA", [
        _paso(CORTE, "Corte", 0),
        _paso(DOBLEZ, "Doblez", 0),
        _paso(ALM_WIP, "Almacen WIP", 9, es_virtual=True),
    ])
    assert _total_completos(c1) == 0, "las 4 de Doblez aun no completan la ruta"
    assert _total_completos(c2) == 9, "las 9 del buffer si terminaron"


def test_total_completos_cuenta_el_ultimo_paso_para_el_pt():
    """El PT no tiene buffer virtual: sus completas esperan en su ultimo paso
    catalogado (tipicamente Embarques)."""
    pt = _comp("PT-1", [
        _paso(CORTE, "Corte", 5),
        _paso(EMBARQUES, "Embarques", 12),
    ], tipo_material=1)
    assert _total_completos(pt) == 12


def test_total_completos_ultimo_paso_aunque_no_sea_embarques_ni_almacen():
    """No todos los componentes terminan en Embarques/Almacen WIP.

    Uno que acaba en Dock Audit (id ficticio 40) debe reportar el WIP de ese
    ultimo paso — una lista fija {13,16} lo dejaria en 0.
    """
    DOCK_AUDIT = 40
    comp = _comp("C1", [
        _paso(CORTE, "Corte", 5),
        _paso(DOCK_AUDIT, "Dock Audit", 8),
    ])
    assert _total_completos(comp) == 8


def test_total_completos_no_confunde_armado_de_kits_inicial():
    """idProceso=16 al INICIO (PT con ARMADO DE KITS) no es "terminado".

    Sumar por id {13,16} contaria las 30 piezas que esperan ENTRAR al armado
    como si estuvieran completas. Lo correcto es el ultimo paso (Embarques=2).
    """
    SOLDADURA = 6
    pt = _comp("KIT-PT", [
        _paso(ALM_WIP, "Armado de kits", 30),   # primer paso REAL, no virtual
        _paso(SOLDADURA, "Soldadura", 4),
        _paso(EMBARQUES, "Embarques", 2),
    ], tipo_material=1)
    assert _total_completos(pt) == 2


def test_total_completos_sin_ruta_es_cero():
    """Sin ruta no hay donde medir lo terminado."""
    comp = _comp("C1", [])
    assert _total_completos(comp) == 0


def test_total_wip_es_el_inventario_fisico():
    """Para una pieza suelta, cantidad bruta; para un kit, unidades ya armadas."""
    comp = _comp("C1", [_paso(CORTE, "Corte", 7)], wip_total=7)
    assert _total_wip(comp) == 7


# ---- Estructura del archivo -------------------------------------------------

def test_una_hoja_por_pt():
    wb = _abrir(exportar_arboles([
        _arbol([_comp("C1", [_paso(CORTE, "Corte", 1)])], pt="PT-A"),
        _arbol([_comp("C2", [_paso(DOBLEZ, "Doblez", 2)])], pt="PT-B"),
    ]))
    assert wb.sheetnames == ["PT-A", "PT-B"]


def test_nombre_de_hoja_se_sanea_y_no_colisiona():
    """Excel prohibe []:*?/\\ y limita a 31 chars; dos PTs no pueden colisionar."""
    wb = _abrir(exportar_arboles([
        _arbol([_comp("C1", [])], pt="PT/A:1"),
        _arbol([_comp("C2", [])], pt="PT/A:1"),
        _arbol([_comp("C3", [])], pt="X" * 40),
    ]))
    assert wb.sheetnames[0] == "PT-A-1"
    assert wb.sheetnames[1] == "PT-A-1~2"      # desambiguado
    assert len(wb.sheetnames[2]) <= 31          # truncado al limite de Excel


def test_encabezado_de_proceso_agrupa_tres_subcolumnas():
    wb = _abrir(exportar_arboles([
        _arbol([_comp("C1", [
            _paso(CORTE, "Corte", 5, disponibles=2, recibidas=3, liberadas=1),
        ])])
    ]))
    ws = wb.active
    # Fila 4 = encabezado del proceso; fila 5 = subcolumnas.
    fila_grupo = [c.value for c in ws[4]]
    fila_sub = [c.value for c in ws[5]]
    assert "Corte" in fila_grupo
    i = fila_grupo.index("Corte")
    assert fila_sub[i : i + 3] == ["Disponible", "Recibido", "Por transferir"]


def test_los_buckets_del_proceso_van_bajo_su_columna():
    wb = _abrir(exportar_arboles([
        _arbol([_comp("C1", [
            _paso(CORTE, "Corte", 5, disponibles=2, recibidas=3, liberadas=1),
        ], wip_total=5, req_neto=95)])
    ]))
    ws = wb.active
    fila_grupo = [c.value for c in ws[4]]
    i = fila_grupo.index("Corte")  # 0-based
    fila = [c.value for c in ws[6]]  # primera fila de datos
    assert fila[i : i + 3] == [2, 3, 1]


def test_componente_que_no_pasa_por_el_proceso_queda_en_blanco():
    """Blanco != 0: 'no aplica' no debe leerse como 'hay cero piezas'."""
    wb = _abrir(exportar_arboles([
        _arbol([
            _comp("C1", [_paso(CORTE, "Corte", 5, disponibles=5)], idComp=1),
            _comp("C2", [_paso(DOBLEZ, "Doblez", 2, disponibles=2)], idComp=2),
        ])
    ]))
    ws = wb.active
    fila_grupo = [c.value for c in ws[4]]
    i_doblez = fila_grupo.index("Doblez")
    # C1 (fila 6) no pasa por Doblez
    assert ws.cell(row=6, column=i_doblez + 1).value is None


def test_columnas_consolidadas_presentes():
    wb = _abrir(exportar_arboles([_arbol([_comp("C1", [_paso(CORTE, "Corte", 1)])])]))
    titulos = [c.value for c in wb.active[4]]
    for esperado in ("Total WIP", "Total completos", "Requerimiento neto"):
        assert esperado in titulos


def _abrir(contenido: bytes):
    return load_workbook(io.BytesIO(contenido))
