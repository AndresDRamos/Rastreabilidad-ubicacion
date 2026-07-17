"""Exportacion de arboles netteados a Excel (.xlsx).

Una hoja por PT. Cada fila es un componente del BOM; cada proceso de la ruta
aporta un grupo de 3 columnas (Disponible / Recibido / Por transferir) bajo un
encabezado combinado, y al final 3 columnas consolidadas.

Reusa `arbol_service.armar_arbol` — el netteo NO se reimplementa aqui.
"""

from __future__ import annotations

import io
import re

import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..domain.modelo import ArbolPT, NodoComponente


log = structlog.get_logger("rbom_api.services.export")


# Procesos que NO fabrican: no transforman la pieza, solo la retienen.
#   13 = Embarques      -> el PT ya termino y espera salir
#   16 = Almacen WIP    -> buffer / ARMADO DE KITS: el intermedio espera al padre
#
# Doble rol en "Total completos": al no fabricar quedan fuera de la ruta
# productiva, y por eso mismo el WIP que espera en ellos es exactamente el que
# ya termino toda su fabricacion. Ver `_total_completos`.
PROCESOS_NO_FABRICACION = frozenset({13, 16})

# --- Estilos -----------------------------------------------------------------
_AZUL = "3B82F6"
_GRIS_CAB = "F5F5F5"
_BORDE_FINO = Side(style="thin", color="D4D4D4")
_BORDE = Border(left=_BORDE_FINO, right=_BORDE_FINO, top=_BORDE_FINO, bottom=_BORDE_FINO)
_CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Columnas de identificacion del componente (antes de los procesos).
_COLS_ID = [
    ("Nivel", 7),
    ("Componente", 20),
    ("Descripcion", 34),
    ("Tipo", 12),
    ("Cant/PT", 9),
    ("Ruta", 30),
]

# Columnas consolidadas (despues de los procesos).
_COLS_FIN = [
    ("Total WIP", 11),
    ("Total completos", 15),
    ("Requerimiento neto", 18),
    ("Req. total (todos los PTs)", 22),
    ("PTs que lo comparten", 30),
]


def exportar_arboles(arboles: list[ArbolPT]) -> bytes:
    """Genera el .xlsx con una hoja por PT. Devuelve los bytes del archivo."""
    wb = Workbook()
    wb.remove(wb.active)  # quita la hoja vacia por defecto

    usados: set[str] = set()
    for arbol in arboles:
        ws = wb.create_sheet(_nombre_hoja(arbol.pt.PT, usados))
        _escribir_hoja(ws, arbol)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _nombre_hoja(clave_pt: str, usados: set[str]) -> str:
    """Nombre de hoja valido para Excel: <=31 chars, sin []:*?/\\ , unico."""
    limpio = re.sub(r"[\[\]:*?/\\]", "-", clave_pt).strip() or "PT"
    base = limpio[:31]
    nombre = base
    i = 2
    while nombre.lower() in usados:
        # Trunca para dejar espacio al sufijo "~N"
        sufijo = f"~{i}"
        nombre = base[: 31 - len(sufijo)] + sufijo
        i += 1
    usados.add(nombre.lower())
    return nombre


def _procesos_del_arbol(arbol: ArbolPT) -> list[tuple[int, str]]:
    """Procesos (reales) que aparecen en las rutas, en orden de fabricacion.

    Ordena por la posicion promedio del proceso en las rutas donde aparece, para
    que las columnas se lean de izquierda (temprano) a derecha (tardio). El
    buffer virtual `Almacen WIP` se omite: no es un paso fisico.
    """
    posiciones: dict[int, list[float]] = {}
    nombres: dict[int, str] = {}
    for comp in arbol.componentes:
        reales = [p for p in comp.ruta if not p.es_virtual]
        total = len(reales)
        for idx, paso in enumerate(reales):
            nombres[paso.idProceso] = paso.proceso
            # Normaliza la posicion a [0,1]: rutas de distinto largo son comparables.
            posiciones.setdefault(paso.idProceso, []).append(
                idx / max(1, total - 1) if total > 1 else 0.0
            )
    return [
        (idp, nombres[idp])
        for idp in sorted(
            posiciones, key=lambda i: sum(posiciones[i]) / len(posiciones[i])
        )
    ]


def _total_wip(comp: NodoComponente) -> float:
    """Piezas del componente que existen fisicamente en WIP.

    Para una pieza suelta es su cantidad bruta. Para un padre/kit son las
    unidades del propio padre ya ensambladas, es decir kits completos en piso.
    En ambos casos es el mismo dato: `wip_total`.
    """
    return comp.wip_total


def _total_completos(comp: NodoComponente) -> float:
    """Piezas que ya terminaron TODA la fabricacion del componente.

    Son las que esperan en los procesos que no fabrican: `Almacen WIP` (el
    buffer donde el intermedio espera a su padre) y `Embarques` (donde el PT
    espera salir). Llegar ahi implica haber pasado toda la ruta.

    La clave esta en la semantica del WIP (trampa #10 del contrato):
    `wip_en_paso[X]` = piezas esperando **entrar** a X, no piezas que ya
    pasaron X. Por eso "completas" no es el WIP del ultimo proceso de
    fabricacion — esas piezas son justo las que aun NO lo han hecho.

    Caso canonico 91711066-RA (verificado contra BD):
      - 90358715-RA: ruta Corte -> Doblez, 4 pzs esperando Doblez -> 0 completas
        (les falta Doblez). El WIP del ultimo proceso real diria 4: falso.
      - 91711040-RA: 9 pzs en Almacen WIP tras Doblez -> 9 completas. El WIP del
        ultimo proceso real diria 0: falso.
    """
    return sum(
        p.wip_en_paso
        for p in comp.ruta
        if p.idProceso in PROCESOS_NO_FABRICACION
    )


def _escribir_hoja(ws: Worksheet, arbol: ArbolPT) -> None:
    procesos = _procesos_del_arbol(arbol)
    n_id = len(_COLS_ID)

    # --- Fila 1: titulo del PT ------------------------------------------------
    total_cols = n_id + len(procesos) * 3 + len(_COLS_FIN)
    ws.cell(row=1, column=1, value=f"{arbol.pt.PT} — {arbol.pt.Descripcion}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, min(total_cols, 8)))

    ws.cell(row=2, column=1, value=(
        f"Cliente: {arbol.pt.Cliente}  ·  Ciudad: {arbol.pt.Ciudad}  ·  "
        f"Demanda: {arbol.pt.PiezasPend:,.0f} pzs  ·  "
        f"Past-due: {arbol.pt.PiezasPastDue:,.0f} pzs"
    ))
    ws.cell(row=2, column=1).font = Font(size=9, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, min(total_cols, 8)))

    # --- Filas 4-5: encabezados de dos niveles --------------------------------
    fila_grupo, fila_sub = 4, 5

    for i, (titulo, ancho) in enumerate(_COLS_ID, start=1):
        _cabecera(ws, fila_grupo, i, titulo, fila_sub)
        ws.column_dimensions[get_column_letter(i)].width = ancho

    col = n_id + 1
    for idp, nombre in procesos:
        # Encabezado combinado del proceso sobre sus 3 subcolumnas.
        ws.merge_cells(start_row=fila_grupo, start_column=col, end_row=fila_grupo, end_column=col + 2)
        c = ws.cell(row=fila_grupo, column=col, value=nombre)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=_AZUL)
        c.alignment = _CENTRO
        c.border = _BORDE
        for j, sub in enumerate(("Disponible", "Recibido", "Por transferir")):
            s = ws.cell(row=fila_sub, column=col + j, value=sub)
            s.font = Font(bold=True, size=8)
            s.fill = PatternFill("solid", fgColor=_GRIS_CAB)
            s.alignment = _CENTRO
            s.border = _BORDE
            ws.column_dimensions[get_column_letter(col + j)].width = 11
        col += 3

    for i, (titulo, ancho) in enumerate(_COLS_FIN):
        _cabecera(ws, fila_grupo, col + i, titulo, fila_sub)
        ws.column_dimensions[get_column_letter(col + i)].width = ancho

    # --- Datos ----------------------------------------------------------------
    idx_proc = {idp: n_id + 1 + i * 3 for i, (idp, _) in enumerate(procesos)}
    fila = fila_sub + 1
    for comp in sorted(arbol.componentes, key=lambda c: (c.nivel, c.clave)):
        pasos = {p.idProceso: p for p in comp.ruta if not p.es_virtual}

        valores = [
            comp.nivel,
            comp.clave,
            comp.descripcion or "",
            "PT" if comp.tipo_material == 1 else "Intermedio",
            comp.cantidad_ensamble_total,
            comp.cadena_ruta,
        ]
        for i, v in enumerate(valores, start=1):
            _celda(ws, fila, i, v)

        for idp, base in idx_proc.items():
            paso = pasos.get(idp)
            if paso is None:
                # El componente no pasa por este proceso: se deja en blanco (no 0),
                # para no confundir "no aplica" con "hay cero piezas".
                for j in range(3):
                    _celda(ws, fila, base + j, None)
                continue
            _celda(ws, fila, base + 0, paso.disponibles)
            _celda(ws, fila, base + 1, paso.recibidas)
            _celda(ws, fila, base + 2, paso.liberadas)  # = "Por transferir"

        ru = comp.req_universo
        finales = [
            _total_wip(comp),
            _total_completos(comp),
            comp.req_neto,
            ru.req_neto_total if ru else None,
            ", ".join(ru.pts) if ru and ru.n_pts > 1 else "",
        ]
        for i, v in enumerate(finales):
            _celda(ws, fila, col + i, v)
        # Resalta el requerimiento neto: es la columna que el planner acciona.
        ws.cell(row=fila, column=col + 2).font = Font(bold=True)

        fila += 1

    # Congela encabezados + columnas de identificacion.
    ws.freeze_panes = ws.cell(row=fila_sub + 1, column=n_id + 1)

    if arbol.advertencias:
        fila += 1
        ws.cell(row=fila, column=1, value="Advertencias").font = Font(bold=True, size=9)
        for adv in arbol.advertencias:
            fila += 1
            ws.cell(row=fila, column=1, value=adv).font = Font(size=8, color="B45309")


def _cabecera(ws: Worksheet, fila: int, col: int, titulo: str, fila_sub: int) -> None:
    """Encabezado simple que abarca las dos filas de cabecera."""
    ws.merge_cells(start_row=fila, start_column=col, end_row=fila_sub, end_column=col)
    c = ws.cell(row=fila, column=col, value=titulo)
    c.font = Font(bold=True, size=9)
    c.fill = PatternFill("solid", fgColor=_GRIS_CAB)
    c.alignment = _CENTRO
    c.border = _BORDE


def _celda(ws: Worksheet, fila: int, col: int, valor) -> None:
    c = ws.cell(row=fila, column=col, value=valor)
    c.border = _BORDE
    if isinstance(valor, (int, float)):
        c.number_format = "#,##0"
        c.alignment = Alignment(horizontal="right")
